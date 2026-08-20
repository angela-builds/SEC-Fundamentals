from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    import yfinance as yf
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing package.")
    print("Please run: pip install pandas openpyxl yfinance")
    raise


BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = BASE_DIR / "outputs" / "excel"


def load_existing_workbook_data(ticker):
    excel_path = EXCEL_DIR / f"{ticker}_annual_fundamentals.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError(f"Cannot find file: {excel_path}")

    per_share_df = pd.read_excel(excel_path, sheet_name="per_share")
    raw_df = pd.read_excel(excel_path, sheet_name="raw_annual_facts")

    return excel_path, per_share_df, raw_df


def get_eps_periods(raw_df):
    eps_df = raw_df[raw_df["metric"] == "eps_diluted"].copy()

    if eps_df.empty:
        raise ValueError("Cannot find eps_diluted in raw_annual_facts.")

    eps_df["start"] = pd.to_datetime(eps_df["start"], errors="coerce")
    eps_df["end"] = pd.to_datetime(eps_df["end"], errors="coerce")

    eps_df = eps_df.dropna(subset=["start", "end"])

    eps_periods = eps_df[["fy", "start", "end"]].drop_duplicates()
    eps_periods = eps_periods.sort_values("fy").reset_index(drop=True)

    return eps_periods


def download_price_history(ticker, start_date, end_date):
    yahoo_ticker = yf.Ticker(ticker)

    price_df = yahoo_ticker.history(
        start=start_date,
        end=end_date,
        auto_adjust=False,
    )

    if price_df.empty:
        raise ValueError(f"No price history found for {ticker}.")

    price_df = price_df.reset_index()
    price_df["Date"] = pd.to_datetime(price_df["Date"]).dt.tz_localize(None)

    return price_df


def get_latest_price(ticker):
    yahoo_ticker = yf.Ticker(ticker)

    recent = yahoo_ticker.history(period="5d", auto_adjust=False)

    if recent.empty:
        raise ValueError(f"No recent price data found for {ticker}.")

    latest_row = recent.dropna(subset=["Close"]).iloc[-1]

    latest_date = latest_row.name
    if hasattr(latest_date, "tz_localize"):
        latest_date = latest_date.tz_localize(None)

    return {
        "current_price_date": latest_date,
        "current_price": latest_row["Close"],
    }


def build_valuation_table(ticker, per_share_df, raw_df):
    eps_periods = get_eps_periods(raw_df)

    base_df = per_share_df[["fy", "eps_diluted"]].copy()
    valuation_df = base_df.merge(eps_periods, on="fy", how="left")

    min_start = valuation_df["start"].min()
    max_end = valuation_df["end"].max()

    if pd.isna(min_start) or pd.isna(max_end):
        raise ValueError("Cannot determine fiscal year start/end dates.")

    price_df = download_price_history(
        ticker=ticker,
        start_date=min_start.strftime("%Y-%m-%d"),
        end_date=(max_end + pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
    )

    rows = []

    for _, row in valuation_df.iterrows():
        fy = int(row["fy"])
        fiscal_start = row["start"]
        fiscal_end = row["end"]
        eps = row["eps_diluted"]

        period_prices = price_df[
            (price_df["Date"] >= fiscal_start)
            & (price_df["Date"] <= fiscal_end)
        ].copy()

        if period_prices.empty or pd.isna(eps) or eps == 0:
            high_close = pd.NA
            high_date = pd.NaT
            high_pe = pd.NA
        else:
            high_idx = period_prices["Close"].idxmax()
            high_row = period_prices.loc[high_idx]

            high_close = high_row["Close"]
            high_date = high_row["Date"]
            high_pe = high_close / eps

        rows.append(
            {
                "fy": fy,
                "fiscal_start": fiscal_start,
                "fiscal_end": fiscal_end,
                "eps_diluted": eps,
                "fiscal_year_high_close": high_close,
                "fiscal_year_high_date": high_date,
                "fiscal_year_high_pe": high_pe,
            }
        )

    result_df = pd.DataFrame(rows)

    latest = get_latest_price(ticker)
    latest_eps = result_df.sort_values("fy").iloc[-1]["eps_diluted"]

    result_df["current_price"] = latest["current_price"]
    result_df["current_price_date"] = latest["current_price_date"]
    result_df["current_pe_using_latest_eps"] = latest["current_price"] / latest_eps
    result_df["pe_fetched_at"] = datetime.now()

    return result_df


def format_valuation_sheet(excel_path):
    wb = load_workbook(excel_path)
    ws = wb["valuation"]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.freeze_panes = "A2"

    money_columns = {
        "eps_diluted",
        "fiscal_year_high_close",
        "current_price",
    }

    date_columns = {
        "fiscal_start",
        "fiscal_end",
        "fiscal_year_high_date",
        "current_price_date",
    }

    datetime_columns = {
        "pe_fetched_at",
    }

    pe_columns = {
        "fiscal_year_high_pe",
        "current_pe_using_latest_eps",
    }

    headers = {cell.value: cell.column for cell in ws[1]}

    for header, column_index in headers.items():
        if header in money_columns:
            for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.number_format = '"$"0.00'

        if header in date_columns:
            for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.number_format = "yyyy-mm-dd"

        if header in datetime_columns:
            for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.number_format = "yyyy-mm-dd hh:mm"

        if header in pe_columns:
            for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.number_format = "0.0x"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 28)

    wb.save(excel_path)


def save_valuation_sheet(excel_path, valuation_df):
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        valuation_df.to_excel(writer, sheet_name="valuation", index=False)

    format_valuation_sheet(excel_path)


def main():
    import sys as _sys
    ticker = _sys.argv[1].strip().upper() if len(_sys.argv) > 1 else input("Enter ticker, for example AAPL or PG: ").strip().upper()

    excel_path, per_share_df, raw_df = load_existing_workbook_data(ticker)

    valuation_df = build_valuation_table(ticker, per_share_df, raw_df)

    save_valuation_sheet(excel_path, valuation_df)

    print(f"Added valuation sheet to: {excel_path}")
    print("\nValuation preview:")
    print(
        valuation_df[
            [
                "fy",
                "eps_diluted",
                "fiscal_year_high_close",
                "fiscal_year_high_pe",
                "current_price",
                "current_pe_using_latest_eps",
                "pe_fetched_at",
            ]
        ].to_string(index=False)
    )


if __name__ == "__main__":
    main()