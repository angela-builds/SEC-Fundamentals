from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing package.")
    print("Please run: pip install pandas openpyxl")
    raise


BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = BASE_DIR / "outputs" / "excel"


METRIC_DEFINITIONS = [
    {
        "metric": "EPS 每股盈餘",
        "module": "Per Share",
        "purpose": "每股賺多少錢",
        "source": "SEC",
        "formula_or_note": "EarningsPerShareDiluted",
    },
    {
        "metric": "Dividend 股息",
        "module": "Per Share",
        "purpose": "股東現金回報",
        "source": "SEC",
        "formula_or_note": "Different companies may use different dividend tags.",
    },
    {
        "metric": "Shares 股數",
        "module": "Per Share",
        "purpose": "稀釋股數、回購判斷",
        "source": "SEC",
        "formula_or_note": "WeightedAverageNumberOfDilutedSharesOutstanding",
    },
    {
        "metric": "Book Value Per Share",
        "module": "Per Share",
        "purpose": "每股淨值",
        "source": "SEC calculated",
        "formula_or_note": "Equity / Shares. Be careful with stock splits.",
    },
    {
        "metric": "FCF 自由現金流",
        "module": "Cash Flow",
        "purpose": "現金創造能力",
        "source": "SEC calculated",
        "formula_or_note": "Operating Cash Flow - CapEx",
    },
    {
        "metric": "Net Margin",
        "module": "Profitability",
        "purpose": "獲利品質",
        "source": "SEC calculated",
        "formula_or_note": "Net Income / Revenue",
    },
    {
        "metric": "ROE",
        "module": "Profitability",
        "purpose": "股東權益報酬率",
        "source": "SEC calculated",
        "formula_or_note": "Net Income / Average Equity",
    },
    {
        "metric": "IC",
        "module": "Balance Sheet",
        "purpose": "Invested Capital，投入資本",
        "source": "SEC calculated",
        "formula_or_note": "Debt + Equity - Cash",
    },
    {
        "metric": "D/E",
        "module": "Balance Sheet",
        "purpose": "負債權益比",
        "source": "SEC calculated",
        "formula_or_note": "Total Debt / Equity",
    },
    {
        "metric": "P/E 本益比",
        "module": "Valuation",
        "purpose": "估值",
        "source": "SEC + yfinance",
        "formula_or_note": "Price / EPS. Historical P/E depends on price date rule.",
    },
]


SUMMARY_METRICS = [
    {
        "metric": "EPS",
        "sheet": "per_share",
        "column": "eps_diluted",
        "format_type": "per_share",
    },
    {
        "metric": "Dividend Per Share",
        "sheet": "per_share",
        "column": "dividend_per_share",
        "format_type": "per_share",
    },
    {
        "metric": "Shares Diluted",
        "sheet": "per_share",
        "column": "shares_diluted",
        "format_type": "shares",
    },
    {
        "metric": "Book Value Per Share",
        "sheet": "per_share",
        "column": "book_value_per_share",
        "format_type": "per_share",
    },
    {
        "metric": "FCF",
        "sheet": "cash_flow",
        "column": "fcf",
        "format_type": "money",
    },
    {
        "metric": "FCF Margin",
        "sheet": "cash_flow",
        "column": "fcf_margin",
        "format_type": "percent",
    },
    {
        "metric": "Net Margin",
        "sheet": "profitability",
        "column": "net_margin",
        "format_type": "percent",
    },
    {
        "metric": "ROE",
        "sheet": "profitability",
        "column": "roe",
        "format_type": "percent",
    },
    {
        "metric": "Invested Capital",
        "sheet": "balance_sheet",
        "column": "invested_capital",
        "format_type": "money",
    },
    {
        "metric": "D/E",
        "sheet": "balance_sheet",
        "column": "debt_to_equity",
        "format_type": "multiple",
    },
    {
        "metric": "Fiscal Year High P/E",
        "sheet": "valuation",
        "column": "fiscal_year_high_pe",
        "format_type": "multiple",
    },
    {
        "metric": "Current P/E",
        "sheet": "valuation",
        "column": "current_pe_using_latest_eps",
        "format_type": "multiple",
    },
]


def load_workbook_sheets(excel_path):
    sheets = {}

    xls = pd.ExcelFile(excel_path)

    for sheet_name in xls.sheet_names:
        sheets[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)

    return sheets


def build_charts_data(sheets):
    required_sheets = [
        "per_share",
        "cash_flow",
        "profitability",
        "balance_sheet",
        "valuation",
    ]

    for sheet_name in required_sheets:
        if sheet_name not in sheets:
            raise ValueError(f"Missing sheet: {sheet_name}")

    charts_df = sheets["per_share"][["fy"]].copy()

    charts_df = charts_df.merge(
        sheets["per_share"][
            [
                "fy",
                "eps_diluted",
                "dividend_per_share",
                "shares_diluted",
                "book_value_per_share",
            ]
        ],
        on="fy",
        how="left",
    )

    charts_df = charts_df.merge(
        sheets["cash_flow"][["fy", "fcf", "fcf_margin"]],
        on="fy",
        how="left",
    )

    charts_df = charts_df.merge(
        sheets["profitability"][["fy", "net_margin", "roe"]],
        on="fy",
        how="left",
    )

    charts_df = charts_df.merge(
        sheets["balance_sheet"][["fy", "invested_capital", "debt_to_equity"]],
        on="fy",
        how="left",
    )

    charts_df = charts_df.merge(
        sheets["valuation"][["fy", "fiscal_year_high_pe", "current_pe_using_latest_eps"]],
        on="fy",
        how="left",
    )

    charts_df = charts_df.sort_values("fy").reset_index(drop=True)

    return charts_df


KNOWN_SPLIT_RATIOS = [1.5, 2, 3, 4, 5, 6, 7, 10, 20, 25, 50]
SPLIT_RATIO_TOLERANCE = 0.06

# Columns whose cross-year comparability breaks if a stock split isn't
# accounted for (per-share dollar amounts and share counts).
SPLIT_SENSITIVE_COLUMNS = {
    "eps_diluted",
    "dividend_per_share",
    "book_value_per_share",
    "shares_diluted",
}


def _has_unresolved_split(per_share_df):
    """True if shares_diluted has a YoY jump close to a known split ratio
    AND it hasn't already been resolved by 08_apply_split_adjustments.py
    (i.e. there's no split_adjusted=True row explaining it)."""
    if "shares_diluted" not in per_share_df.columns or "fy" not in per_share_df.columns:
        return False

    df = per_share_df.sort_values("fy").reset_index(drop=True)
    already_adjusted = bool(df.get("split_adjusted", pd.Series(dtype=bool)).any())
    if already_adjusted:
        return False

    multiples = df["shares_diluted"] / df["shares_diluted"].shift(1)
    for multiple in multiples.dropna():
        closest = min(KNOWN_SPLIT_RATIOS, key=lambda r: abs(r - multiple))
        if closest > 1.2 and abs(closest - multiple) / closest <= SPLIT_RATIO_TOLERANCE:
            return True
    return False


def build_summary(sheets):
    rows = []

    unresolved_split = "per_share" in sheets and _has_unresolved_split(sheets["per_share"])

    for item in SUMMARY_METRICS:
        sheet_name = item["sheet"]
        column = item["column"]

        if sheet_name not in sheets:
            continue

        df = sheets[sheet_name].copy()

        if column not in df.columns or "fy" not in df.columns:
            continue

        df = df[["fy", column]].dropna()

        if df.empty:
            continue

        df = df.sort_values("fy")
        latest_row = df.iloc[-1]

        if unresolved_split and column in SPLIT_SENSITIVE_COLUMNS:
            rows.append(
                {
                    "metric": item["metric"],
                    "source_sheet": sheet_name,
                    "latest_year": int(latest_row["fy"]),
                    "latest_value": latest_row[column],
                    "10y_avg": "N/A - possible stock split, see raw_annual_facts",
                    "10y_high": "N/A - possible stock split, see raw_annual_facts",
                    "10y_low": "N/A - possible stock split, see raw_annual_facts",
                    "years_covered": df["fy"].nunique(),
                    "format_type": "text_flag",
                }
            )
            continue

        years_covered = df["fy"].nunique()
        rows.append(
            {
                "metric": item["metric"],
                "source_sheet": sheet_name,
                "latest_year": int(latest_row["fy"]),
                "latest_value": latest_row[column],
                "10y_avg": df[column].mean(),
                "10y_high": df[column].max(),
                "10y_low": df[column].min(),
                "years_covered": years_covered,
                "format_type": item["format_type"],
            }
        )

    summary_df = pd.DataFrame(rows)

    if not summary_df.empty and (summary_df["years_covered"] < 10).any():
        min_years = int(summary_df["years_covered"].min())
        summary_df = summary_df.rename(
            columns={"10y_avg": f"avg (up to {min_years}y, see years_covered col)"}
        )
        print(
            f"  [note] This workbook has less than 10 years of history for at least one "
            f"metric (as few as {min_years} years). Columns labeled '10y_high'/'10y_low' "
            f"and the renamed avg column reflect the 'years_covered' column per row, not "
            f"a fixed 10-year window."
        )

    return summary_df


def build_metric_definition():
    return pd.DataFrame(METRIC_DEFINITIONS)


def format_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    top_alignment = Alignment(vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = top_alignment

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 36)


def apply_number_formats(wb):
    number_format_by_type = {
        "money": '"$"#,##0',
        "percent": "0.0%",
        "per_share": '"$"0.00',
        "shares": "#,##0",
        "multiple": "0.0x",
    }

    # Summary uses row-level format_type.
    if "summary" in wb.sheetnames:
        ws = wb["summary"]
        headers = {cell.value: cell.column for cell in ws[1]}

        value_columns = [
            "latest_value",
            "10y_avg",
            "10y_high",
            "10y_low",
        ]

        format_col = headers.get("format_type")

        if format_col:
            for row_idx in range(2, ws.max_row + 1):
                format_type = ws.cell(row=row_idx, column=format_col).value
                number_format = number_format_by_type.get(format_type)

                if not number_format:
                    continue

                for col_name in value_columns:
                    col_idx = headers.get(col_name)
                    if col_idx:
                        ws.cell(row=row_idx, column=col_idx).number_format = number_format

    # Charts data uses column-level format.
    if "charts_data" in wb.sheetnames:
        ws = wb["charts_data"]
        headers = {cell.value: cell.column for cell in ws[1]}

        format_map = {
            "eps_diluted": "per_share",
            "dividend_per_share": "per_share",
            "shares_diluted": "shares",
            "book_value_per_share": "per_share",
            "fcf": "money",
            "fcf_margin": "percent",
            "net_margin": "percent",
            "roe": "percent",
            "invested_capital": "money",
            "debt_to_equity": "multiple",
            "fiscal_year_high_pe": "multiple",
            "current_pe_using_latest_eps": "multiple",
        }

        for header, format_type in format_map.items():
            col_idx = headers.get(header)
            number_format = number_format_by_type.get(format_type)

            if not col_idx or not number_format:
                continue

            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format


def save_summary_outputs(excel_path, summary_df, charts_df, definition_df):
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        charts_df.to_excel(writer, sheet_name="charts_data", index=False)
        definition_df.to_excel(writer, sheet_name="metric_definition", index=False)

    wb = load_workbook(excel_path)

    preferred_order = [
        "summary",
        "charts_data",
        "metric_definition",
        "per_share",
        "cash_flow",
        "profitability",
        "balance_sheet",
        "valuation",
        "raw_annual_facts",
        "data_notes",
    ]

    existing = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [existing[name] for name in preferred_order if name in existing]

    for sheet_name in ["summary", "charts_data", "metric_definition"]:
        if sheet_name in wb.sheetnames:
            format_sheet(wb[sheet_name])

    apply_number_formats(wb)

    wb.save(excel_path)


def main():
    import sys as _sys
    ticker = _sys.argv[1].strip().upper() if len(_sys.argv) > 1 else input("Enter ticker, for example AAPL or PG: ").strip().upper()

    excel_path = EXCEL_DIR / f"{ticker}_annual_fundamentals.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError(f"Cannot find file: {excel_path}")

    sheets = load_workbook_sheets(excel_path)

    charts_df = build_charts_data(sheets)
    summary_df = build_summary(sheets)
    definition_df = build_metric_definition()

    save_summary_outputs(excel_path, summary_df, charts_df, definition_df)

    print(f"Added summary, charts_data, and metric_definition sheets to: {excel_path}")
    print("\nSummary preview:")
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()