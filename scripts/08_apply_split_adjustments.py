"""
08_apply_split_adjustments.py

Reads config/confirmed_splits.csv (human-confirmed splits only -- see
07_detect_stock_splits.py) and, for each confirmed split:

  1. Keeps the original as-reported per-share numbers in new columns
     (never deletes/overwrites raw data).
  2. Adds split-adjusted columns for fiscal years BEFORE the split, so the
     whole 10-year series is on one consistent basis:
       - eps_diluted, dividend_per_share, book_value_per_share  -> divided by ratio
       - shares_diluted                                          -> multiplied by ratio
  3. Adds a "split_adjusted" flag column so it's visible which rows were touched.
  4. Appends a note to data_notes explaining what was done and why.
  5. Rebuilds summary / charts_data using the adjusted numbers, so 10y_high /
     10y_low / 10y_avg are no longer mixing pre- and post-split bases.

Run 07_detect_stock_splits.py first and manually confirm each candidate
before adding it to config/confirmed_splits.csv. This script trusts that
file completely -- it does not re-verify splits itself.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = BASE_DIR / "outputs" / "excel"
CONFIG_PATH = BASE_DIR / "config" / "confirmed_splits.csv"

ADJUSTED_PER_SHARE_COLS = ["eps_diluted", "dividend_per_share", "book_value_per_share"]
ADJUSTED_SHARE_COUNT_COLS = ["shares_diluted"]


def _load_module(script_name: str):
    """04_build_summary.py starts with a digit, so it can't be `import`ed
    normally -- load it by file path instead."""
    path = BASE_DIR / "scripts" / script_name
    spec = importlib.util.spec_from_file_location(script_name.replace(".py", ""), path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def adjust_per_share_multi(per_share: pd.DataFrame, splits: list[dict]) -> pd.DataFrame:
    """Apply ALL confirmed splits for one ticker in a single pass, using a
    cumulative adjustment factor per fiscal year. This matters whenever a
    ticker has more than one split (e.g. NVDA: 4:1 in FY2022, 10:1 in
    FY2025) -- years before BOTH splits need to be divided by 4*10=40, not
    just by the nearer split's ratio, or the older years stay wrong.
    """
    per_share = per_share.sort_values("fy").reset_index(drop=True)

    for col in ADJUSTED_PER_SHARE_COLS + ADJUSTED_SHARE_COUNT_COLS:
        as_reported_col = f"{col}_as_reported"
        if as_reported_col not in per_share.columns:
            per_share[as_reported_col] = per_share[col]

    if "split_adjusted" not in per_share.columns:
        per_share["split_adjusted"] = False

    # Cumulative factor for each fiscal year = product of ratios for every
    # confirmed split whose fy_effective is AFTER that year.
    def cumulative_factor(fy: int) -> float:
        factor = 1.0
        for s in splits:
            if fy < s["fy_effective"]:
                factor *= s["ratio"]
        return factor

    per_share["_cum_factor"] = per_share["fy"].apply(cumulative_factor)

    for col in ADJUSTED_PER_SHARE_COLS:
        per_share[col] = per_share[f"{col}_as_reported"] / per_share["_cum_factor"]

    for col in ADJUSTED_SHARE_COUNT_COLS:
        per_share[col] = per_share[f"{col}_as_reported"] * per_share["_cum_factor"]

    per_share["split_adjusted"] = per_share["_cum_factor"] != 1.0
    per_share = per_share.drop(columns=["_cum_factor"])

    return per_share


def adjust_per_share(per_share: pd.DataFrame, fy_effective: int, ratio: float) -> pd.DataFrame:
    """Kept for backward compatibility / single-split cases. Prefer
    adjust_per_share_multi when a ticker may have more than one split."""
    return adjust_per_share_multi(per_share, [{"fy_effective": fy_effective, "ratio": ratio}])


def append_data_note(sheets: dict, ticker: str, fy_effective: int, ratio: float, split_date: str, note: str):
    if "data_notes" not in sheets:
        return sheets

    new_row = pd.DataFrame(
        [
            {
                "fy": fy_effective,
                "metric": "shares_diluted / eps_diluted / dividend_per_share / book_value_per_share",
                "fact_type": "split_adjustment",
                "sec_tag": "N/A",
                "unit": "N/A",
                "period_days": None,
                "form": "N/A",
                "filed": split_date,
                "accn": f"MANUAL: {note} (ratio {ratio}:1, applied to FY before {fy_effective})",
            }
        ]
    )
    if sheets["data_notes"].empty:
        sheets["data_notes"] = new_row
    else:
        existing = sheets["data_notes"].dropna(axis=1, how="all")
        sheets["data_notes"] = pd.concat([existing, new_row], ignore_index=True)
    return sheets


def process_ticker(ticker: str, splits: list[dict], build_summary_mod):
    excel_path = EXCEL_DIR / f"{ticker}_annual_fundamentals.xlsx"
    if not excel_path.exists():
        print(f"  [{ticker}] file not found, skipping: {excel_path}")
        return

    sheets = build_summary_mod.load_workbook_sheets(excel_path)

    if "per_share" not in sheets:
        print(f"  [{ticker}] no per_share sheet, skipping")
        return

    # 04-08 assume the full 01-06 pipeline already ran successfully for this
    # ticker (so valuation/cash_flow/profitability/balance_sheet all exist).
    # Fail loudly with an actionable message instead of a raw stack trace if
    # something's missing -- this usually means 03_add_valuation.py (or one
    # of the other steps) hasn't actually completed for this ticker yet.
    required_sheets = {"cash_flow", "profitability", "balance_sheet", "valuation", "raw_annual_facts"}
    missing = required_sheets - set(sheets.keys())
    if missing:
        print(f"  [{ticker}] SKIPPED -- missing sheet(s): {', '.join(sorted(missing))}")
        print(f"           This usually means the 01-05 pipeline didn't fully complete for {ticker}.")
        print(f"           Fix: python scripts\\run_pipeline.py {ticker} --skip-fetch")
        print(f"           (or rerun the specific missing step, e.g. 03_add_valuation.py {ticker})")
        return

    sheets["per_share"] = adjust_per_share_multi(sheets["per_share"], splits)

    for s in splits:
        sheets = append_data_note(
            sheets, ticker, s["fy_effective"], s["ratio"], s.get("split_date", ""), s.get("note", "")
        )

    # Write per_share and data_notes back first.
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        sheets["per_share"].to_excel(writer, sheet_name="per_share", index=False)
        if "data_notes" in sheets:
            sheets["data_notes"].to_excel(writer, sheet_name="data_notes", index=False)

    # Reload (so summary/charts pick up the adjusted per_share values) and rebuild.
    sheets = build_summary_mod.load_workbook_sheets(excel_path)
    charts_df = build_summary_mod.build_charts_data(sheets)
    summary_df = build_summary_mod.build_summary(sheets)
    definition_df = build_summary_mod.build_metric_definition()

    build_summary_mod.save_summary_outputs(excel_path, summary_df, charts_df, definition_df)

    ratios_desc = ", ".join(f"{s['ratio']}:1 @FY{s['fy_effective']}" for s in splits)
    print(f"  [{ticker}] applied cumulative adjustment ({ratios_desc}), rebuilt summary/charts_data")


def main():
    if not CONFIG_PATH.exists():
        print(f"No confirmed splits file at {CONFIG_PATH}. Nothing to do.")
        return

    confirmed = pd.read_csv(CONFIG_PATH)
    if confirmed.empty:
        print("confirmed_splits.csv is empty. Nothing to do.")
        return

    build_summary_mod = _load_module("04_build_summary.py")

    confirmed["ticker"] = confirmed["ticker"].str.strip().str.upper()

    for ticker, group in confirmed.groupby("ticker"):
        splits = [
            {
                "fy_effective": int(row["split_fy_effective"]),
                "ratio": float(row["ratio"]),
                "split_date": str(row.get("split_date", "")),
                "note": str(row.get("note", "")),
            }
            for _, row in group.iterrows()
        ]
        print(f"Processing {ticker} ({len(splits)} confirmed split(s))...")
        process_ticker(ticker=ticker, splits=splits, build_summary_mod=build_summary_mod)

    print("\nDone. Raw as-reported numbers are preserved in *_as_reported columns")
    print("in the per_share sheet -- nothing was deleted, only added/adjusted.")


if __name__ == "__main__":
    main()
