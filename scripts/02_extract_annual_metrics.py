import json
from pathlib import Path

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing package: pandas or openpyxl")
    print("Please run: pip install pandas openpyxl")
    raise


BASE_DIR = Path(__file__).resolve().parents[1]
RAW_DIR = BASE_DIR / "data" / "raw"
EXCEL_DIR = BASE_DIR / "outputs" / "excel"

ANALYSIS_YEARS = 10
MIN_YEARS = 3  # below this, the "trend" is too short to be meaningful at all


FLOW_METRIC_TAGS = {
    "revenue": [
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "Revenues",
        "SalesRevenueNet",
    ],
    "net_income": [
        "NetIncomeLoss",
    ],
    "eps_diluted": [
        "EarningsPerShareDiluted",
    ],
    "shares_diluted": [
        "WeightedAverageNumberOfDilutedSharesOutstanding",
    ],
    "operating_cash_flow": [
        "NetCashProvidedByUsedInOperatingActivities",
        "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations",
    ],
    "capex": [
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "PaymentsToAcquireProductiveAssets",
        "CapitalExpenditures",
    ],
    "dividends_paid": [
        "PaymentsOfDividends",
        "PaymentsOfDividendsCommonStock",
        "PaymentsOfOrdinaryDividends",
    ],
}


POINT_METRIC_TAGS = {
    "equity": [
        "StockholdersEquity",
        "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
    ],
    "cash": [
        "CashAndCashEquivalentsAtCarryingValue",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
    ],
    "short_term_debt": [
        "ShortTermBorrowings",
        "ShortTermDebt",
    ],
    "long_term_debt_current": [
        "LongTermDebtCurrent",
    ],
    "long_term_debt_noncurrent": [
        "LongTermDebtNoncurrent",
    ],
    "long_term_debt_total": [
        "LongTermDebt",
    ],
}


UNIT_BY_METRIC = {
    "revenue": "USD",
    "net_income": "USD",
    "eps_diluted": "USD/shares",
    "shares_diluted": "shares",
    "operating_cash_flow": "USD",
    "capex": "USD",
    "dividends_paid": "USD",
    "equity": "USD",
    "cash": "USD",
    "short_term_debt": "USD",
    "long_term_debt_current": "USD",
    "long_term_debt_noncurrent": "USD",
    "long_term_debt_total": "USD",
}


def load_companyfacts(ticker):
    path = RAW_DIR / f"{ticker}_companyfacts.json"

    if not path.exists():
        raise FileNotFoundError(f"Cannot find file: {path}")

    with open(path, mode="r", encoding="utf-8") as file:
        return json.load(file)


def get_facts_for_tag(companyfacts, tag, unit):
    facts = companyfacts.get("facts", {}).get("us-gaap", {})

    if tag not in facts:
        return []

    units = facts[tag].get("units", {})

    if unit not in units:
        return []

    return units[unit]


def calculate_period_days(start, end):
    if not start or not end:
        return None

    start_dt = pd.to_datetime(start, errors="coerce")
    end_dt = pd.to_datetime(end, errors="coerce")

    if pd.isna(start_dt) or pd.isna(end_dt):
        return None

    return (end_dt - start_dt).days


def select_annual_flow_facts(facts):
    """
    Select annual flow facts.

    IMPORTANT:
    Do NOT use the SEC CompanyFacts `fy` field as the fiscal year.

    Some SEC CompanyFacts records can contain comparative annual facts
    from earlier fiscal periods while carrying the FY value of the later
    filing.

    Example:
        start = 2021-02-01
        end   = 2022-01-30
        SEC fy = 2024

    The actual fiscal period belongs to FY2022, so we use the year of
    the fiscal-period end date instead:

        end = 2022-01-30 -> fy = 2022

    This is particularly important for NVDA CapEx data.
    """

    rows = []

    for item in facts:
        form = item.get("form")
        fp = item.get("fp")
        val = item.get("val")
        start = item.get("start")
        end = item.get("end")
        filed = item.get("filed")

        if form not in ["10-K", "10-K/A"]:
            continue

        if fp != "FY":
            continue

        if val is None:
            continue

        if not end:
            continue

        period_days = calculate_period_days(start, end)

        if period_days is None:
            continue

        # Flow metrics should cover roughly one fiscal year.
        if not (300 <= period_days <= 450):
            continue

        end_dt = pd.to_datetime(end, errors="coerce")

        if pd.isna(end_dt):
            continue

        # ------------------------------------------------------------
        # IMPORTANT FIX:
        # Use the actual fiscal-period END year instead of SEC `fy`.
        # ------------------------------------------------------------
        fiscal_year = int(end_dt.year)

        rows.append(
            {
                "fy": fiscal_year,
                "start": start,
                "end": end,
                "period_days": period_days,
                "val": val,
                "form": form,
                "filed": filed,
                "accn": item.get("accn"),
                "fact_type": "flow",
            }
        )

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    df["filed_dt"] = pd.to_datetime(
        df["filed"],
        errors="coerce",
    )

    # If multiple filings contain the same fiscal-year period,
    # keep the latest filed version.
    df = df.sort_values(
        ["fy", "filed_dt"],
        ascending=[True, True],
    )

    df = df.drop_duplicates(
        subset=["fy"],
        keep="last",
    )

    df = df.drop(
        columns=["filed_dt"]
    )

    return df


def select_annual_point_facts(facts):
    rows = []

    for item in facts:
        form = item.get("form")
        fp = item.get("fp")
        fy = item.get("fy")
        val = item.get("val")
        end = item.get("end")
        filed = item.get("filed")

        if form not in ["10-K", "10-K/A"]:
            continue

        if fp != "FY":
            continue

        if fy is None or val is None or end is None:
            continue

        rows.append(
            {
                "fy": int(fy),
                "start": item.get("start"),
                "end": end,
                "period_days": None,
                "val": val,
                "form": form,
                "filed": filed,
                "accn": item.get("accn"),
                "fact_type": "point",
            }
        )

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    df["filed_dt"] = pd.to_datetime(
        df["filed"],
        errors="coerce",
    )

    df["end_dt"] = pd.to_datetime(
        df["end"],
        errors="coerce",
    )

    # For point-in-time metrics, use the latest end date
    # and latest filed value for each fiscal year.
    df = df.sort_values(
        ["fy", "end_dt", "filed_dt"],
        ascending=[True, False, False],
    )

    df = df.drop_duplicates(
        subset=["fy"],
        keep="first",
    )

    df = df.drop(
        columns=["filed_dt", "end_dt"]
    )

    return df


def extract_metric(companyfacts, metric_name, metric_type):
    if metric_type == "flow":
        candidate_tags = FLOW_METRIC_TAGS[metric_name]
    else:
        candidate_tags = POINT_METRIC_TAGS[metric_name]

    unit = UNIT_BY_METRIC[metric_name]
    frames = []

    for tag_priority, tag in enumerate(candidate_tags):

        facts = get_facts_for_tag(
            companyfacts,
            tag,
            unit,
        )

        if metric_type == "flow":
            annual_df = select_annual_flow_facts(facts)
        else:
            annual_df = select_annual_point_facts(facts)

        if annual_df.empty:
            continue

        annual_df = annual_df[
            [
                "fy",
                "start",
                "end",
                "period_days",
                "val",
                "form",
                "filed",
                "accn",
                "fact_type",
            ]
        ].copy()

        annual_df["metric"] = metric_name
        annual_df["sec_tag"] = tag
        annual_df["unit"] = unit
        annual_df["tag_priority"] = tag_priority

        frames.append(annual_df)

    if not frames:
        return pd.DataFrame(
            columns=[
                "fy",
                "start",
                "end",
                "period_days",
                "val",
                "form",
                "filed",
                "accn",
                "fact_type",
                "metric",
                "sec_tag",
                "unit",
            ]
        )

    combined_df = pd.concat(
        frames,
        ignore_index=True,
    )

    combined_df["filed_dt"] = pd.to_datetime(
        combined_df["filed"],
        errors="coerce",
    )

    combined_df = combined_df.sort_values(
        [
            "fy",
            "tag_priority",
            "filed_dt",
        ],
        ascending=[
            True,
            True,
            False,
        ],
    )

    selected_df = combined_df.drop_duplicates(
        subset=["fy"],
        keep="first",
    )

    selected_df = selected_df.drop(
        columns=[
            "tag_priority",
            "filed_dt",
        ]
    )

    selected_df = selected_df.sort_values(
        "fy"
    ).reset_index(
        drop=True
    )

    return selected_df


def build_annual_table(
    companyfacts,
    analysis_years=ANALYSIS_YEARS,
):
    metric_frames = []

    for metric_name in FLOW_METRIC_TAGS:
        metric_frames.append(
            extract_metric(
                companyfacts,
                metric_name,
                "flow",
            )
        )

    for metric_name in POINT_METRIC_TAGS:
        metric_frames.append(
            extract_metric(
                companyfacts,
                metric_name,
                "point",
            )
        )

    raw_df = pd.concat(
        metric_frames,
        ignore_index=True,
    )

    if raw_df.empty:
        raise ValueError(
            "No annual facts found. Please check the SEC JSON file."
        )

    available_years = sorted(
        raw_df["fy"]
        .dropna()
        .unique()
    )

    if len(available_years) < MIN_YEARS:
        raise ValueError(
            f"Only found {len(available_years)} fiscal years. "
            f"Need at least {MIN_YEARS} years for a meaningful trend."
        )

    if len(available_years) < analysis_years:
        print(
            f"  [WARNING] Only {len(available_years)} fiscal years "
            f"of data available "
            f"(wanted {analysis_years}). Proceeding with what's there -- "
            f"this ticker likely IPO'd or was spun off recently. "
            f"Any '10y_avg/high/low' labels downstream should be read as "
            f"covering {len(available_years)} years, not 10."
        )

    required_years = (
        available_years[-(analysis_years + 1):]
        if len(available_years) > analysis_years
        else available_years
    )

    analysis_year_list = (
        required_years[-analysis_years:]
        if len(required_years) > analysis_years
        else required_years
    )

    raw_df = raw_df[
        raw_df["fy"].isin(required_years)
    ].copy()

    pivot_df = raw_df.pivot_table(
        index="fy",
        columns="metric",
        values="val",
        aggfunc="first",
    ).reset_index()

    pivot_df = pivot_df.sort_values(
        "fy"
    ).reset_index(
        drop=True
    )

    return (
        raw_df,
        pivot_df,
        analysis_year_list,
    )


def safe_divide(numerator, denominator):
    return numerator / denominator.replace(
        {0: pd.NA}
    )


def calculate_metrics(
    pivot_df,
    analysis_year_list,
):
    df = pivot_df.copy()

    required_columns = (
        list(FLOW_METRIC_TAGS.keys())
        + list(POINT_METRIC_TAGS.keys())
    )

    for column in required_columns:
        if column not in df.columns:
            df[column] = pd.NA

    df["total_debt_parts"] = (
        df["short_term_debt"].fillna(0)
        + df["long_term_debt_current"].fillna(0)
        + df["long_term_debt_noncurrent"].fillna(0)
    )

    df["total_debt"] = df["total_debt_parts"]

    df.loc[
        df["total_debt"] == 0,
        "total_debt",
    ] = df["long_term_debt_total"]

    df["invested_capital"] = (
        df["total_debt"].fillna(0)
        + df["equity"].fillna(0)
        - df["cash"].fillna(0)
    )

    df["fcf"] = (
        df["operating_cash_flow"]
        - df["capex"].abs()
    )

    df["fcf_margin"] = safe_divide(
        df["fcf"],
        df["revenue"],
    )

    df["book_value_per_share"] = safe_divide(
        df["equity"],
        df["shares_diluted"],
    )

    df["dividend_per_share"] = safe_divide(
        df["dividends_paid"].abs(),
        df["shares_diluted"],
    )

    df["net_margin"] = safe_divide(
        df["net_income"],
        df["revenue"],
    )

    df["average_equity"] = (
        df["equity"]
        + df["equity"].shift(1)
    ) / 2

    df["roe"] = safe_divide(
        df["net_income"],
        df["average_equity"],
    )

    df["debt_to_equity"] = safe_divide(
        df["total_debt"],
        df["equity"],
    )

    df["revenue_growth_yoy"] = (
        df["revenue"].pct_change()
    )

    df["eps_growth_yoy"] = (
        df["eps_diluted"].pct_change()
    )

    df["fcf_growth_yoy"] = (
        df["fcf"].pct_change()
    )

    analysis_df = df[
        df["fy"].isin(analysis_year_list)
    ].copy()

    per_share_df = analysis_df[
        [
            "fy",
            "eps_diluted",
            "shares_diluted",
            "dividends_paid",
            "dividend_per_share",
            "book_value_per_share",
        ]
    ].copy()

    cash_flow_df = analysis_df[
        [
            "fy",
            "operating_cash_flow",
            "capex",
            "fcf",
            "fcf_margin",
            "fcf_growth_yoy",
        ]
    ].copy()

    profitability_df = analysis_df[
        [
            "fy",
            "revenue",
            "revenue_growth_yoy",
            "net_income",
            "net_margin",
            "roe",
        ]
    ].copy()

    balance_sheet_df = analysis_df[
        [
            "fy",
            "equity",
            "cash",
            "total_debt",
            "invested_capital",
            "debt_to_equity",
        ]
    ].copy()

    return (
        per_share_df,
        cash_flow_df,
        profitability_df,
        balance_sheet_df,
    )


def build_data_notes(raw_df):
    notes = (
        raw_df[
            [
                "fy",
                "metric",
                "fact_type",
                "sec_tag",
                "unit",
                "period_days",
                "form",
                "filed",
                "accn",
            ]
        ]
        .drop_duplicates()
        .sort_values(
            [
                "metric",
                "fy",
            ]
        )
        .reset_index(
            drop=True
        )
    )

    return notes


def format_worksheet(ws):
    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True
    )

    center_alignment = Alignment(
        horizontal="center"
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = (
            column_cells[0].column_letter
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(
                max_length,
                len(str(value)),
            )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            30,
        )


def apply_number_formats(ws):
    money_columns = {
        "revenue",
        "net_income",
        "operating_cash_flow",
        "capex",
        "fcf",
        "equity",
        "cash",
        "total_debt",
        "invested_capital",
        "dividends_paid",
    }

    percent_columns = {
        "revenue_growth_yoy",
        "eps_growth_yoy",
        "fcf_growth_yoy",
        "fcf_margin",
        "net_margin",
        "roe",
        "debt_to_equity",
    }

    per_share_columns = {
        "eps_diluted",
        "dividend_per_share",
        "book_value_per_share",
    }

    share_columns = {
        "shares_diluted",
    }

    headers = {
        cell.value: cell.column
        for cell in ws[1]
    }

    for header, column_index in headers.items():

        if header in money_columns:
            for row in ws.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
            ):
                for cell in row:
                    cell.number_format = '"$"#,##0'

        if header in percent_columns:
            for row in ws.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
            ):
                for cell in row:
                    cell.number_format = "0.0%"

        if header in per_share_columns:
            for row in ws.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
            ):
                for cell in row:
                    cell.number_format = '"$"0.00'

        if header in share_columns:
            for row in ws.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
            ):
                for cell in row:
                    cell.number_format = "#,##0"


def save_to_excel(
    ticker,
    raw_df,
    per_share_df,
    cash_flow_df,
    profitability_df,
    balance_sheet_df,
    notes_df,
):
    EXCEL_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        EXCEL_DIR
        / f"{ticker}_annual_fundamentals.xlsx"
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        per_share_df.to_excel(
            writer,
            sheet_name="per_share",
            index=False,
        )

        cash_flow_df.to_excel(
            writer,
            sheet_name="cash_flow",
            index=False,
        )

        profitability_df.to_excel(
            writer,
            sheet_name="profitability",
            index=False,
        )

        balance_sheet_df.to_excel(
            writer,
            sheet_name="balance_sheet",
            index=False,
        )

        raw_df.to_excel(
            writer,
            sheet_name="raw_annual_facts",
            index=False,
        )

        notes_df.to_excel(
            writer,
            sheet_name="data_notes",
            index=False,
        )

        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            format_worksheet(ws)
            apply_number_formats(ws)

    print(
        f"Saved Excel: {output_path}"
    )


def main():
    import sys as _sys

    ticker = (
        _sys.argv[1].strip().upper()
        if len(_sys.argv) > 1
        else input(
            "Enter ticker, for example AAPL or PG: "
        ).strip().upper()
    )

    companyfacts = load_companyfacts(
        ticker
    )

    (
        raw_df,
        pivot_df,
        analysis_year_list,
    ) = build_annual_table(
        companyfacts
    )

    (
        per_share_df,
        cash_flow_df,
        profitability_df,
        balance_sheet_df,
    ) = calculate_metrics(
        pivot_df,
        analysis_year_list,
    )

    notes_df = build_data_notes(
        raw_df
    )

    print(
        "Analysis years extracted:"
    )

    print(
        ", ".join(
            str(year)
            for year in sorted(
                analysis_year_list
            )
        )
    )

    print(
        "\nRaw years kept for YoY calculation:"
    )

    print(
        ", ".join(
            str(year)
            for year in sorted(
                pivot_df["fy"].tolist()
            )
        )
    )

    print(
        "\nKey metrics preview:"
    )

    preview = profitability_df[
        [
            "fy",
            "revenue",
            "net_income",
            "net_margin",
            "roe",
        ]
    ]

    print(
        preview.to_string(
            index=False
        )
    )

    save_to_excel(
        ticker,
        raw_df,
        per_share_df,
        cash_flow_df,
        profitability_df,
        balance_sheet_df,
        notes_df,
    )


if __name__ == "__main__":
    main()
