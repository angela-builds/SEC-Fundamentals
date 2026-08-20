from pathlib import Path

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing package.")
    print("Please run: pip install pandas openpyxl matplotlib")
    raise


BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = BASE_DIR / "outputs" / "excel"
CHARTS_DIR = BASE_DIR / "outputs" / "charts"

SINGLE_CHARTS = [
    {
        "column": "eps_diluted",
        "title": "EPS",
        "ylabel": "USD per share",
        "filename": "01_eps.png",
        "money": True,
    },
    {
        "column": "dividend_per_share",
        "title": "Dividend Per Share",
        "ylabel": "USD per share",
        "filename": "02_dividend_per_share.png",
        "money": True,
    },
    {
        "column": "shares_diluted",
        "title": "Shares Diluted",
        "ylabel": "Shares",
        "filename": "03_shares_diluted.png",
        "money": False,
    },
    {
        "column": "book_value_per_share",
        "title": "Book Value Per Share",
        "ylabel": "USD per share",
        "filename": "04_book_value_per_share.png",
        "money": True,
    },
    {
        "column": "fcf",
        "title": "Free Cash Flow",
        "ylabel": "USD",
        "filename": "05_fcf.png",
        "money": False,
    },
    {
        "column": "fcf_margin",
        "title": "FCF Margin",
        "ylabel": "%",
        "filename": "06_fcf_margin.png",
        "percent": True,
    },
    {
        "column": "net_margin",
        "title": "Net Margin",
        "ylabel": "%",
        "filename": "07_net_margin.png",
        "percent": True,
    },
]


def load_charts_data(ticker):
    excel_path = EXCEL_DIR / f"{ticker}_annual_fundamentals.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError(f"Cannot find file: {excel_path}")

    charts_df = pd.read_excel(excel_path, sheet_name="charts_data")
    charts_df = charts_df.sort_values("fy").reset_index(drop=True)

    return excel_path, charts_df


def format_large_number(value):
    abs_value = abs(value)

    if abs_value >= 1_000_000_000:
        return f"{value / 1_000_000_000:.1f}B"

    if abs_value >= 1_000_000:
        return f"{value / 1_000_000:.1f}M"

    return f"{value:,.0f}"


def plot_single_chart(df, ticker, chart_config, output_dir):
    column = chart_config["column"]
    title = chart_config["title"]
    ylabel = chart_config["ylabel"]
    output_path = output_dir / chart_config["filename"]

    if column not in df.columns:
        print(f"Skip missing column: {column}")
        return None

    plot_df = df[["fy", column]].dropna().copy()

    if plot_df.empty:
        print(f"Skip empty column: {column}")
        return None

    y_values = plot_df[column]

    if chart_config.get("percent"):
        y_values = y_values * 100

    fig, ax = plt.subplots(figsize=(8, 4.2))

    ax.plot(plot_df["fy"], y_values, marker="o", linewidth=2)

    ax.set_title(f"{ticker} - {title}", fontsize=14, weight="bold")
    ax.set_xlabel("Fiscal Year")
    ax.set_ylabel(ylabel)
    ax.grid(True, alpha=0.3)

    ax.set_xticks(plot_df["fy"])
    ax.tick_params(axis="x", rotation=45)

    if chart_config.get("percent"):
        ax.yaxis.set_major_formatter(lambda value, _: f"{value:.0f}%")

    if column == "shares_diluted":
        ax.yaxis.set_major_formatter(lambda value, _: format_large_number(value))

    if column == "fcf":
        ax.yaxis.set_major_formatter(lambda value, _: f"${format_large_number(value)}")

    if chart_config.get("money") and column != "fcf":
        ax.yaxis.set_major_formatter(lambda value, _: f"${value:.2f}")

    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)

    return output_path


def plot_pe_chart(df, ticker, output_dir):
    output_path = output_dir / "08_pe_valuation.png"

    required_columns = [
        "fy",
        "fiscal_year_high_pe",
        "current_pe_using_latest_eps",
    ]

    for column in required_columns:
        if column not in df.columns:
            print(f"Skip P/E chart. Missing column: {column}")
            return None

    plot_df = df[required_columns].dropna().copy()

    if plot_df.empty:
        print("Skip P/E chart. No data.")
        return None

    fig, ax = plt.subplots(figsize=(8, 4.2))

    ax.plot(
        plot_df["fy"],
        plot_df["fiscal_year_high_pe"],
        marker="o",
        linewidth=2,
        label="Fiscal Year High P/E",
    )

    ax.plot(
        plot_df["fy"],
        plot_df["current_pe_using_latest_eps"],
        marker="o",
        linewidth=2,
        linestyle="--",
        label="Current P/E",
    )

    ax.set_title(f"{ticker} - P/E Valuation", fontsize=14, weight="bold")
    ax.set_xlabel("Fiscal Year")
    ax.set_ylabel("P/E")
    ax.grid(True, alpha=0.3)
    ax.legend()

    ax.set_xticks(plot_df["fy"])
    ax.tick_params(axis="x", rotation=45)
    ax.yaxis.set_major_formatter(lambda value, _: f"{value:.1f}x")

    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)

    return output_path


def build_chart_images(ticker, charts_df):
    output_dir = CHARTS_DIR / ticker
    output_dir.mkdir(parents=True, exist_ok=True)

    image_paths = []

    for chart_config in SINGLE_CHARTS:
        image_path = plot_single_chart(charts_df, ticker, chart_config, output_dir)

        if image_path:
            image_paths.append(image_path)

    pe_image_path = plot_pe_chart(charts_df, ticker, output_dir)

    if pe_image_path:
        image_paths.append(pe_image_path)

    return image_paths


def remove_existing_charts_sheet(wb):
    if "charts" in wb.sheetnames:
        ws = wb["charts"]
        wb.remove(ws)


def insert_charts_into_excel(excel_path, image_paths):
    wb = load_workbook(excel_path)

    remove_existing_charts_sheet(wb)

    ws = wb.create_sheet("charts", 1)

    ws["A1"] = "Charts"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"].alignment = Alignment(horizontal="center")

    positions = [
        "A3",
        "J3",
        "A25",
        "J25",
        "A47",
        "J47",
        "A69",
        "J69",
    ]

    for image_path, position in zip(image_paths, positions):
        img = Image(str(image_path))
        img.width = 640
        img.height = 336
        ws.add_image(img, position)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["J"].width = 12

    wb.save(excel_path)


def main():
    import sys as _sys
    ticker = _sys.argv[1].strip().upper() if len(_sys.argv) > 1 else input("Enter ticker, for example AAPL or PG: ").strip().upper()

    excel_path, charts_df = load_charts_data(ticker)

    image_paths = build_chart_images(ticker, charts_df)

    if not image_paths:
        print("No charts generated.")
        return

    insert_charts_into_excel(excel_path, image_paths)

    print(f"Generated {len(image_paths)} charts.")
    print(f"Charts saved under: {CHARTS_DIR / ticker}")
    print(f"Inserted charts into: {excel_path}")


if __name__ == "__main__":
    main()